import os
import json
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import List, Dict, Any

class ExcelProcessor:
    def __init__(self):
        self.default_headers = [
            'Page',
            'Type Document',
            'Date',
            'Montant',
            'Devise',
            'Emetteur',
            'Destinataire',
            'Numéro Document',
            'Autres Informations',
            'Image Path',
            'Date Traitement'
        ]
    
    def create_template_excel(self, output_path: str):
        """
        Crée un fichier Excel template avec les colonnes par défaut
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Résultats Analyse"
        
        # Ajouter les en-têtes
        for col, header in enumerate(self.default_headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Ajuster la largeur des colonnes
        for col in range(1, len(self.default_headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        wb.save(output_path)
        return output_path
    
    def update_excel_with_results(self, excel_path: str, results: List[Dict[str, Any]], output_path: str):
        """
        Met à jour le fichier Excel avec les résultats d'analyse
        """
        try:
            # Charger le workbook existant ou créer un nouveau
            if os.path.exists(excel_path):
                wb = load_workbook(excel_path)
                ws = wb.active
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = "Résultats Analyse"
                
                # Ajouter les en-têtes si le fichier est nouveau
                for col, header in enumerate(self.default_headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
            
            # Trouver la prochaine ligne disponible
            next_row = ws.max_row + 1
            
            # Ajouter les résultats
            for result in results:
                page_num = result.get('page', 0)
                ai_result = result.get('ai_result', {})
                image_path = result.get('image_path', '')
                
                # Extraire les données de l'AI
                content = ai_result.get('content', {})
                if isinstance(content, str):
                    # Si c'est une string, essayer de la parser
                    try:
                        content = json.loads(content)
                    except:
                        content = {'raw_text': content}
                
                # Préparer les données pour Excel
                row_data = [
                    page_num,
                    content.get('type_document', ''),
                    content.get('date', ''),
                    content.get('montant', ''),
                    content.get('devise', ''),
                    content.get('emetteur', ''),
                    content.get('destinataire', ''),
                    content.get('numero_document', ''),
                    json.dumps(content.get('autres_informations', {}), ensure_ascii=False),
                    image_path,
                    datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                ]
                
                # Ajouter la ligne
                for col, value in enumerate(row_data, 1):
                    ws.cell(row=next_row, column=col, value=value)
                
                next_row += 1
            
            # Ajuster la largeur des colonnes
            for col in range(1, len(self.default_headers) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 15
            
            # Sauvegarder
            wb.save(output_path)
            return output_path
            
        except Exception as e:
            raise Exception(f"Erreur lors de la mise à jour Excel: {str(e)}")
    
    def read_excel_data(self, excel_path: str) -> List[Dict[str, Any]]:
        """
        Lit les données d'un fichier Excel
        """
        try:
            wb = load_workbook(excel_path, data_only=True)
            ws = wb.active
            
            data = []
            headers = []
            
            # Lire les en-têtes
            for col in range(1, ws.max_column + 1):
                header = ws.cell(row=1, column=col).value
                headers.append(header)
            
            # Lire les données
            for row in range(2, ws.max_row + 1):
                row_data = {}
                for col in range(1, ws.max_column + 1):
                    value = ws.cell(row=row, column=col).value
                    row_data[headers[col-1]] = value
                data.append(row_data)
            
            return data
            
        except Exception as e:
            raise Exception(f"Erreur lors de la lecture Excel: {str(e)}")
    
    def add_custom_headers(self, excel_path: str, custom_headers: List[str]):
        """
        Ajoute des en-têtes personnalisés au fichier Excel
        """
        try:
            wb = load_workbook(excel_path)
            ws = wb.active
            
            # Ajouter les nouveaux en-têtes
            start_col = ws.max_column + 1
            for col, header in enumerate(custom_headers, start_col):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            wb.save(excel_path)
            
        except Exception as e:
            raise Exception(f"Erreur lors de l'ajout d'en-têtes: {str(e)}")
    
    def format_excel_file(self, excel_path: str):
        """
        Applique un formatage au fichier Excel
        """
        try:
            wb = load_workbook(excel_path)
            ws = wb.active
            
            # Formater les en-têtes
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Formater les données
            for row in range(2, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    
                    # Alterner les couleurs de fond
                    if row % 2 == 0:
                        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            
            # Ajuster la largeur des colonnes
            for col in range(1, ws.max_column + 1):
                ws.column_dimensions[get_column_letter(col)].width = 15
            
            wb.save(excel_path)
            
        except Exception as e:
            raise Exception(f"Erreur lors du formatage: {str(e)}") 